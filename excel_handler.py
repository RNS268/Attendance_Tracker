"""
Excel File Handler
Manages attendance records in Excel format
"""

import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from config import Config

def get_current_attendance_file_path():
    """
    Returns the full path to today's (or current week's) attendance Excel file
    """
    from datetime import datetime
    import os
    from config import Config

    week_num = datetime.now().isocalendar()[1]
    year = datetime.now().year
    filename = Config.EXCEL_FILENAME_FORMAT.format(week=week_num, year=year)
    return os.path.join(Config.ATTENDANCE_FOLDER, filename)


def get_week_number(date):
    """
    Calculate ISO week number from date
    
    Args:
        date: datetime object
    
    Returns:
        Tuple: (week_number, year)
    """
    iso_calendar = date.isocalendar()
    week_number = iso_calendar[1]  # ISO week number (1-53)
    year = iso_calendar[0]         # ISO year
    return week_number, year


def get_excel_filename(date):
    """
    Generate Excel filename for given date
    
    Args:
        date: datetime object
    
    Returns:
        String: full path to Excel file
    """
    week_num, year = get_week_number(date)
    
    # Create filename using format from config
    filename = Config.EXCEL_FILENAME_FORMAT.format(week=week_num, year=year)
    # Example: "Attendance_Week_52_2025.xlsx"
    
    # Create full path
    filepath = os.path.join(Config.ATTENDANCE_FOLDER, filename)
    
    return filepath


def create_excel_file(filepath):
    """
    Create new Excel file with headers
    
    Args:
        filepath: string (full path to create file)
    
    Returns:
        Boolean: True if successful, False otherwise
    """
    try:
        # Create new workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"
        
        # Add headers
        headers = ["Date", "Time", "Student ID", "Student Name"]
        sheet.append(headers)
        
        # Bold the headers
        for cell in sheet[1]:  # Row 1 is header row
            cell.font = Font(bold=True)
        
        # Set column widths
        sheet.column_dimensions['A'].width = 12  # Date
        sheet.column_dimensions['B'].width = 10  # Time
        sheet.column_dimensions['C'].width = 15  # Student ID
        sheet.column_dimensions['D'].width = 25  # Student Name
        
        # Save file
        workbook.save(filepath)
        
        print(f"✓ Created new Excel file: {os.path.basename(filepath)}")
        return True
        
    except Exception as e:
        print(f"✗ Error creating Excel file: {e}")
        return False


def check_duplicate_entry(filepath, student_id, date_str):
    """
    Check if student already marked attendance today
    
    Args:
        filepath: string (path to Excel file)
        student_id: string (e.g., "24054-EC-001")
        date_str: string (e.g., "30-12-2025")
    
    Returns:
        Boolean: True if duplicate exists, False otherwise
    """
    try:
        # If file doesn't exist, no duplicate possible
        if not os.path.exists(filepath):
            return False
        
        # Load workbook
        workbook = load_workbook(filepath)
        sheet = workbook.active
        
        # Check all rows (skip header row 1)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3:  # Skip empty rows
                continue
            
            row_date = str(row[0])  # Date column
            row_id = str(row[2])    # ID column
            
            # Compare date and ID
            if row_date == date_str and row_id == student_id:
                return True  # Duplicate found!
        
        return False  # No duplicate
        
    except Exception as e:
        print(f"✗ Error checking duplicate: {e}")
        return False  # On error, assume no duplicate (safer)


def add_attendance_entry(student_id, student_name):
    """
    Add new attendance entry to Excel file
    
    Args:
        student_id: string (e.g., "24054-EC-001")
        student_name: string (e.g., "Rajesh Kumar")
    
    Returns:
        Tuple: (success: boolean, message: string)
    """
    try:
        # Get current date and time
        current_datetime = datetime.now()
        date_str = current_datetime.strftime(Config.DATE_FORMAT)
        time_str = current_datetime.strftime(Config.TIME_FORMAT)
        
        # Get Excel filename for this week
        filepath = get_excel_filename(current_datetime)
        
        # Ensure attendance_files folder exists
        os.makedirs(Config.ATTENDANCE_FOLDER, exist_ok=True)
        
        # Create file if it doesn't exist
        if not os.path.exists(filepath):
            success = create_excel_file(filepath)
            if not success:
                return False, "Failed to create Excel file"
        
        # Check for duplicate entry
        if check_duplicate_entry(filepath, student_id, date_str):
            return False, f"{student_name} already marked attendance today"
        
        # Add new entry
        workbook = load_workbook(filepath)
        sheet = workbook.active
        
        # Append new row: [Date, Time, ID, Name]
        new_row = [date_str, time_str, student_id, student_name]
        sheet.append(new_row)
        
        # Save file
        workbook.save(filepath)
        
        # Print success message
        print(f"✓ Attendance marked for {student_name} at {time_str}")
        
        return True, f"Attendance marked successfully for {student_name}"
        
    except PermissionError:
        return False, "Excel file is open in another program. Please close it and try again."
    except Exception as e:
        print(f"✗ Error adding attendance: {e}")
        return False, f"Error: {str(e)}"


def get_today_attendance_count(date=None):
    """
    Count how many students marked attendance today
    
    Args:
        date: datetime object (optional, defaults to today)
    
    Returns:
        Integer: number of attendance entries today
    """
    try:
        if date is None:
            date = datetime.now()
        
        filepath = get_excel_filename(date)
        
        # If file doesn't exist, count is 0
        if not os.path.exists(filepath):
            return 0
        
        # Load file and count entries for today
        workbook = load_workbook(filepath)
        sheet = workbook.active
        
        date_str = date.strftime(Config.DATE_FORMAT)
        
        count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 1:  # Skip empty rows
                continue
            
            if str(row[0]) == date_str:
                count += 1
        
        return count
        
    except Exception as e:
        print(f"✗ Error counting today's attendance: {e}")
        return 0


def get_today_attendance_list(date=None):
    """
    Get list of students who marked attendance today
    
    Args:
        date: datetime object (optional, defaults to today)
    
    Returns:
        List: [{"name": "...", "id": "...", "time": "..."}, ...]
    """
    try:
        if date is None:
            date = datetime.now()
        
        filepath = get_excel_filename(date)
        
        # If file doesn't exist, return empty list
        if not os.path.exists(filepath):
            return []
        
        # Load file and get today's entries
        workbook = load_workbook(filepath)
        sheet = workbook.active
        
        date_str = date.strftime(Config.DATE_FORMAT)
        
        attendance_list = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 4:  # Skip empty/incomplete rows
                continue
            
            row_date = str(row[0])
            if row_date == date_str:
                attendance_list.append({
                    "name": str(row[3]),  # Student Name
                    "id": str(row[2]),    # Student ID
                    "time": str(row[1])   # Time
                })
        
        return attendance_list
        
    except Exception as e:
        print(f"✗ Error getting today's attendance list: {e}")
        return []


def get_week_attendance_report(week_number=None, year=None):
    """
    Get attendance report for a specific week
    
    Args:
        week_number: int (optional, defaults to current week)
        year: int (optional, defaults to current year)
    
    Returns:
        List: All attendance entries for the week
    """
    try:
        if week_number is None or year is None:
            current_date = datetime.now()
            week_number, year = get_week_number(current_date)
        
        filename = Config.EXCEL_FILENAME_FORMAT.format(week=week_number, year=year)
        filepath = os.path.join(Config.ATTENDANCE_FOLDER, filename)
        
        if not os.path.exists(filepath):
            return []
        
        workbook = load_workbook(filepath)
        sheet = workbook.active
        
        report = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 4:
                continue
            
            report.append({
                "date": str(row[0]),
                "time": str(row[1]),
                "id": str(row[2]),
                "name": str(row[3])
            })
        
        return report
        
    except Exception as e:
        print(f"✗ Error getting week report: {e}")
        return []


def clear_today_attendance():
    """
    Clear today's attendance entries from the Excel file
    
    Returns:
        Tuple: (success: boolean, message: string)
    """
    try:
        from openpyxl import load_workbook
        
        current_date = datetime.now()
        filepath = get_excel_filename(current_date)
        
        if not os.path.exists(filepath):
            return False, "Attendance file does not exist yet for this week."
            
        workbook = load_workbook(filepath)
        sheet = workbook.active
        
        date_str = current_date.strftime(Config.DATE_FORMAT)
        
        # We need to collect rows to delete first to avoid index shifting issues
        rows_to_keep = []
        header = None
        
        # Get header
        for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            header = row
            break
            
        if not header:
            return False, "File is empty or corrupted."
            
        rows_to_keep.append(header)
        
        cleared_count = 0
        
        # Check all other rows
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 1:
                continue
                
            row_date = str(row[0])
            if row_date != date_str:
                rows_to_keep.append(row)
            else:
                cleared_count += 1
        
        if cleared_count == 0:
            return True, "No attendance records found for today to clear."
            
        # Create a new workbook to replace the old one (cleaner than deleting rows)
        new_workbook = Workbook()
        new_sheet = new_workbook.active
        new_sheet.title = "Attendance"
        
        # Append kept rows
        for row in rows_to_keep:
            new_sheet.append(row)
            
        # Re-apply styles
        for cell in new_sheet[1]:  # Header row
            cell.font = Font(bold=True)
            
        new_sheet.column_dimensions['A'].width = 12
        new_sheet.column_dimensions['B'].width = 10
        new_sheet.column_dimensions['C'].width = 15
        new_sheet.column_dimensions['D'].width = 25
        
        # Save
        new_workbook.save(filepath)
        print(f"✓ Cleared {cleared_count} entries for today.")
        
        return True, f"Successfully cleared {cleared_count} entries for today."
        
    except PermissionError:
        return False, "Excel file is open. Please close it and try again."
    except Exception as e:
        print(f"✗ Error clearing attendance: {e}")
        return False, f"Error: {str(e)}"


# Test code
if __name__ == "__main__":
    print("🧪 Testing excel_handler.py...")
    print("=" * 50)
    
    # Test 1: Week number calculation
    print("\n1. Testing week number calculation...")
    test_date = datetime(2025, 12, 30)
    week_num, year = get_week_number(test_date)
    print(f"   Date: {test_date.strftime('%d-%m-%Y')}")
    print(f"   Week: {week_num}, Year: {year}")
    
    # Test 2: Filename generation
    print("\n2. Testing filename generation...")
    filename = get_excel_filename(test_date)
    print(f"   Generated path: {filename}")
    
    # Test 3: Folder creation
    print("\n3. Testing folder creation...")
    os.makedirs(Config.ATTENDANCE_FOLDER, exist_ok=True)
    print(f"   ✓ Folder exists: {os.path.exists(Config.ATTENDANCE_FOLDER)}")
    
    # Test 4: Adding attendance entry
    print("\n4. Testing attendance entry...")
    success, message = add_attendance_entry("24054-EC-001", "Test Student")
    print(f"   Result: {message}")
    
    # Test 5: Duplicate check
    print("\n5. Testing duplicate detection...")
    success2, message2 = add_attendance_entry("24054-EC-001", "Test Student")
    print(f"   Duplicate attempt: {message2}")
    
    # Test 6: Count today's attendance
    print("\n6. Testing attendance count...")
    count = get_today_attendance_count()
    print(f"   Today's count: {count}")
    
    # Test 7: Get today's list
    print("\n7. Testing attendance list...")
    attendance_list = get_today_attendance_list()
    print(f"   Today's attendance: {len(attendance_list)} students")
    for entry in attendance_list:
        print(f"     - {entry['name']} ({entry['id']}) at {entry['time']}")
        
    # Test 8: Clear today's attendance
    print("\n8. Testing clear today's attendance...")
    # Add a dummy entry first to ensure there's something to clear
    add_attendance_entry("DUMMY-001", "Dummy Student")
    success3, message3 = clear_today_attendance()
    print(f"   Result: {message3}")
    
    # Verify count is 0
    new_count = get_today_attendance_count()
    print(f"   Count after clearing: {new_count}")
     
    print("\n" + "=" * 50)
    print("✅ excel_handler.py test complete!")
    
    # Cleanup: Delete test file if it exists
    test_file = get_excel_filename(datetime.now())
    if os.path.exists(test_file):
        try:
            os.remove(test_file)
            print(f"   Cleanup: Removed test file")
        except:
            pass